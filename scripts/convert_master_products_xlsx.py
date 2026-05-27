#!/usr/bin/env python3
"""
Convert master_สินค้า.xlsx → CSV in the exact 12-col template the new
inventory-csv.php importer expects (Thai headers, UTF-8 BOM for Excel).

Usage:
    python scripts/convert_master_products_xlsx.py [INPUT.xlsx] [OUTPUT.csv]

Defaults:
    INPUT  = C:\\Users\\Administrator\\Downloads\\master_สินค้า.xlsx
    OUTPUT = C:\\Users\\Administrator\\Downloads\\master_สินค้า.csv
"""
import sys
import csv
from pathlib import Path
from openpyxl import load_workbook

DEFAULT_IN  = Path(r"C:\Users\Administrator\Downloads\master_สินค้า.xlsx")
DEFAULT_OUT = Path(r"C:\Users\Administrator\Downloads\master_สินค้า.csv")


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_IN
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not src.exists():
        print(f"ERROR: source not found: {src}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filename=str(src), read_only=True, data_only=True)
    ws = wb.active

    rows_out = 0
    # UTF-8 BOM helps Excel auto-detect encoding when reopening the CSV
    with open(dst, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Keep all 12 columns, blank None → ""
            cells = ["" if v is None else str(v).strip() for v in row]
            # Trim trailing all-empty rows
            if i > 1 and not any(cells):
                continue
            w.writerow(cells)
            rows_out += 1

    print(f"OK: wrote {rows_out} rows -> {dst}")


if __name__ == "__main__":
    main()
