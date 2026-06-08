#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate master_products v2 seed SQL from MASTER-DATA-REYA.xlsx (5,572 rows, all UOM)."""
import csv, sys
from openpyxl import load_workbook
from pathlib import Path

SRC = Path(r"C:\Users\Administrator\Downloads\MASTER-DATA-REYA.xlsx")
OUT = Path(r"C:\Users\Administrator\clinicya\database\master_products_v2_seed.sql")

# col indexes
SKU, NAME, NAME_EN, GENERIC, CATEGORY = 0, 1, 2, 3, 4
PRODUCT_CODE = 8
BARCODE, UOM, DESC, USAGE, WARNING = 7, 9, 10, 11, 12
PRICE, SALE, COST = 13, 14, 15
IMG, DISTRIBUTOR, MFR = 28, 29, 30

BATCH = 200


def parse_sku_unit(composite, product_code, uom):
    """
    composite e.g. '4182-กล่อง[50ชิ้น]'  product_code '4182'  uom 'กล่อง[50ชิ้น]'
    Returns (clean_sku, clean_unit, pack_size).
      clean_sku  = product code part (before first '-'); prefer the numeric prefix
      clean_unit = unit before '[' (e.g. กล่อง)
      pack_size  = inside [...] (e.g. 50ชิ้น)
    """
    comp = (str(composite).strip() if composite is not None else "")
    # SKU: take code before first '-' in composite; fall back to Product Code col
    if "-" in comp:
        code = comp.split("-", 1)[0].strip()
    else:
        code = comp
    if not code and product_code is not None:
        code = str(product_code).strip()
    # unit / pack: parse from UOM col (cleaner) else from composite remainder
    src = (str(uom).strip() if uom is not None and str(uom).strip() else "")
    if not src and "-" in comp:
        src = comp.split("-", 1)[1].strip()
    unit, pack = src, ""
    if "[" in src and "]" in src:
        unit = src[:src.index("[")].strip()
        pack = src[src.index("[") + 1: src.rindex("]")].strip()
    return code, unit, pack


def s(v):
    if v is None:
        return "NULL"
    t = str(v).strip()
    if t == "":
        return "NULL"
    t = (t.replace("\\", "\\\\").replace("'", "\\'").replace("\x00", "")
          .replace("\r", " ").replace("\n", " "))
    return "'" + t + "'"


def num(v):
    if v is None:
        return "NULL"
    t = str(v).strip().replace(",", "")
    if t == "":
        return "NULL"
    try:
        return str(float(t))
    except ValueError:
        return "NULL"


def main():
    wb = load_workbook(filename=str(SRC), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    next(it)  # header
    rows = []
    seen = set()
    for r in it:
        if not r or r[SKU] is None or str(r[SKU]).strip() == "":
            continue
        if r[NAME] is None or str(r[NAME]).strip() == "":
            continue
        sku = str(r[SKU]).strip()
        if sku in seen:   # safety dedup on composite SKU
            continue
        seen.add(sku)
        rows.append(r)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as f:
        f.write("-- master_products v2 from MASTER-DATA-REYA.xlsx\n")
        f.write("-- rows: %d\n" % len(rows))
        f.write("SET NAMES utf8mb4;\n")
        f.write("START TRANSACTION;\n")
        f.write("TRUNCATE TABLE master_products;\n\n")
        cols = ("(`uom_ref`,`sku`,`barcode`,`name`,`name_en`,`manufacturer`,`distributor`,"
                "`generic_name`,`category`,`unit`,`pack_size`,`price`,`sale_price`,"
                "`cost_price`,`usage_instructions`,`warning`,`description`,`image_url`,`source`)")
        for i in range(0, len(rows), BATCH):
            chunk = rows[i:i + BATCH]
            f.write("INSERT INTO `master_products` %s VALUES\n" % cols)
            vals = []
            for r in chunk:
                uom_ref = str(r[SKU]).strip()                       # composite, unique
                code, unit, pack = parse_sku_unit(r[SKU], r[PRODUCT_CODE], r[UOM])
                vals.append("(" + ",".join([
                    s(uom_ref), s(code), s(r[BARCODE]), s(r[NAME]), s(r[NAME_EN]), s(r[MFR]),
                    s(r[DISTRIBUTOR]), s(r[GENERIC]), s(r[CATEGORY]), s(unit), s(pack),
                    num(r[PRICE]), num(r[SALE]), num(r[COST]), s(r[USAGE]), s(r[WARNING]),
                    s(r[DESC]), s(r[IMG]), "'cny'",
                ]) + ")")
            f.write(",\n".join(vals) + ";\n\n")
        f.write("COMMIT;\n")
    print("OK rows=%d -> %s (%.1f KB)" % (len(rows), OUT, OUT.stat().st_size / 1024))


if __name__ == "__main__":
    main()
