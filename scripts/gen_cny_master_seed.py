#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build master_products seed from CNYhermesss.xlsx.
One row per UOM (sheet 02_หน่วยนับ) joined with product info (sheet 01_สินค้า by รหัสสินค้า).
image_url set when database/master_img/{code}.jpg exists."""
import os
import openpyxl

SRC = r"C:\Users\Administrator\Downloads\CNYhermesss.xlsx"
IMGDIR = r"C:\Users\Administrator\clinicya\database\master_img"
OUT = r"C:\Users\Administrator\clinicya\database\cny_master_seed.sql"
IMG_BASE = "https://re-ya.com/uploads/master/"
BATCH = 300

# sheet01 (product) col indexes
P_CODE, P_SKU, P_NAME, P_GENERIC, P_EN, P_SPEC = 2, 3, 4, 5, 6, 7
P_CAT, P_DESC, P_USAGE, P_WARN, P_MFR, P_DIST = 9, 14, 15, 16, 17, 18
# sheet02 (uom) col indexes
U_CODE, U_NAME, U_UNIT, U_BARCODE, U_COST, U_PRICE, U_SALE = 1, 3, 5, 7, 8, 9, 10


def sq(v):
    if v is None:
        return "NULL"
    s = str(v).strip()
    if s == "" or s.lower() == "none":
        return "NULL"
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def num(v):
    if v is None or str(v).strip() == "":
        return "NULL"
    try:
        return str(round(float(str(v).replace(",", "").strip()), 2))
    except Exception:
        return "NULL"


def split_unit(unitname):
    s = (str(unitname).strip() if unitname is not None else "")
    if "[" in s and "]" in s:
        return s[:s.index("[")].strip(), s[s.index("[") + 1:s.rindex("]")].strip()
    return s, ""


# images present
imgset = set()
if os.path.isdir(IMGDIR):
    for fn in os.listdir(IMGDIR):
        imgset.add(os.path.splitext(fn)[0])

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# build product info map by code
prod = {}
ws1 = wb["01_สินค้า"]
for r in ws1.iter_rows(min_row=2, values_only=True):
    code = (str(r[P_CODE]).strip() if r[P_CODE] is not None else "")
    if not code:
        continue
    generic = r[P_GENERIC]
    if generic and str(generic).strip() in ("ไม่ระบุ/ตรวจสอบฉลาก", "ไม่ระบุ"):
        generic = r[P_SPEC] or None
    prod[code] = dict(
        name=r[P_NAME], name_en=r[P_EN], generic=generic, category=r[P_CAT],
        desc=r[P_DESC], usage=r[P_USAGE], warn=r[P_WARN], mfr=r[P_MFR], dist=r[P_DIST],
    )

cols = ("(`uom_ref`,`sku`,`barcode`,`name`,`name_en`,`manufacturer`,`distributor`,"
        "`generic_name`,`category`,`unit`,`pack_size`,`price`,`sale_price`,`cost_price`,"
        "`usage_instructions`,`warning`,`description`,`image_url`,`source`,`is_active`)")

seen = set()
rows_sql = []
ws2 = wb["02_หน่วยนับ"]
for r in ws2.iter_rows(min_row=2, values_only=True):
    code = (str(r[U_CODE]).strip() if r[U_CODE] is not None else "")
    if not code:
        continue
    unitname = r[U_UNIT]
    if unitname is None or str(unitname).strip() == "":
        continue
    uom_ref = code + "-" + str(unitname).strip()
    if uom_ref in seen:
        continue
    seen.add(uom_ref)
    unit, pack = split_unit(unitname)
    p = prod.get(code, {})
    name = p.get("name") or r[U_NAME]
    img = IMG_BASE + code + ".jpg" if code in imgset else None
    vals = ",".join([
        sq(uom_ref), sq(code), sq(r[U_BARCODE]), sq(name), sq(p.get("name_en")),
        sq(p.get("mfr")), sq(p.get("dist")), sq(p.get("generic")), sq(p.get("category")),
        sq(unit), sq(pack), num(r[U_PRICE]), num(r[U_SALE]), num(r[U_COST]),
        sq(p.get("usage")), sq(p.get("warn")), sq(p.get("desc")), sq(img), "'cny'", "1",
    ])
    rows_sql.append("(" + vals + ")")

with open(OUT, "w", encoding="utf-8") as f:
    f.write("SET NAMES utf8mb4;\n")
    f.write("DELETE FROM `master_products`;\n")
    for i in range(0, len(rows_sql), BATCH):
        chunk = rows_sql[i:i + BATCH]
        f.write("INSERT INTO `master_products` %s VALUES\n" % cols)
        f.write(",\n".join(chunk))
        f.write(";\n")

with_img = sum(1 for s in seen if s.split("-", 1)[0] in imgset)
print("UOM rows=%d  distinct products=%d  rows_with_image=%d  images_on_disk=%d"
      % (len(rows_sql), len(set(s.split('-', 1)[0] for s in seen)), with_img, len(imgset)))
print("OUT:", OUT)
